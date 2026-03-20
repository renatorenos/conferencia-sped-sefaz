"""
Exportação do resultado da conciliação para .xlsx com 3 abas:
  - Resumo
  - Entradas_Fora_SEFAZ
  - Saidas_Fora_SPED
"""
from __future__ import annotations
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .parsers.sped_parser import EmpresaSped
from .conciliacao import ResultadoConciliacao


_AZUL_HEADER = "1F4E79"
_AZUL_CLARO = "D6E4F0"
_LARANJA_HEADER = "7B3F00"
_LARANJA_CLARO = "FAE5D3"
_CINZA_CLARO = "F2F2F2"


def _header_style(cell, cor_hex: str):
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=cor_hex)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _ajustar_colunas(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)


def _aba_resumo(wb, empresa: EmpresaSped, resultado: ResultadoConciliacao):
    ws = wb.active
    ws.title = "Resumo"

    dados = [
        ("Empresa", empresa.nome),
        ("CNPJ", empresa.cnpj),
        ("UF", empresa.uf),
        ("Período", f"{empresa.dt_ini} a {empresa.dt_fin}"),
        ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("", ""),
        ("SPED — Total entradas lidas", resultado.total_entradas_sped),
        ("SPED — Total saídas lidas", resultado.total_saidas_sped),
        ("SEFAZ — Total entradas lidas", resultado.total_entradas_sefaz),
        ("SEFAZ — Total saídas lidas", resultado.total_saidas_sefaz),
        ("", ""),
        ("Divergência A — Entradas no SEFAZ fora do SPED", len(resultado.entradas_sefaz_fora_sped)),
        ("Divergência B — Saídas no SEFAZ fora do SPED", len(resultado.saidas_sefaz_fora_sped)),
    ]

    for r, (label, valor) in enumerate(dados, start=1):
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 2, valor)

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 30


def _aba_entradas(wb, resultado: ResultadoConciliacao):
    ws = wb.create_sheet("Entradas_Fora_SPED")

    cabecalho = [
        "Chave NF-e", "Número", "Série", "Data Emissão",
        "Emitente", "CNPJ Emitente", "Destinatário", "CNPJ Destinatário",
        "Valor Total", "Situação SEFAZ",
    ]
    for c, col in enumerate(cabecalho, 1):
        cell = ws.cell(1, c, col)
        _header_style(cell, _AZUL_HEADER)

    ws.row_dimensions[1].height = 30

    for r, nota in enumerate(resultado.entradas_sefaz_fora_sped, start=2):
        fill = PatternFill("solid", fgColor=_AZUL_CLARO) if r % 2 == 0 else PatternFill("solid", fgColor=_CINZA_CLARO)
        valores = [
            nota.chave, nota.num_doc, nota.serie, nota.dt_emissao,
            nota.nome_emit, nota.cnpj_emit,
            nota.nome_dest, nota.cnpj_dest,
            nota.vl_total, nota.situacao,
        ]
        for c, v in enumerate(valores, 1):
            cell = ws.cell(r, c, v)
            cell.fill = fill

    _ajustar_colunas(ws)


def _aba_saidas(wb, resultado: ResultadoConciliacao):
    ws = wb.create_sheet("Saidas_Fora_SPED")

    cabecalho = [
        "Chave NF-e", "Número", "Série", "Data Emissão",
        "Emitente", "CNPJ Emitente", "Destinatário", "CNPJ Destinatário",
        "Valor Total", "Situação SEFAZ",
    ]
    for c, col in enumerate(cabecalho, 1):
        cell = ws.cell(1, c, col)
        _header_style(cell, _LARANJA_HEADER)

    ws.row_dimensions[1].height = 30

    for r, nota in enumerate(resultado.saidas_sefaz_fora_sped, start=2):
        fill = PatternFill("solid", fgColor=_LARANJA_CLARO) if r % 2 == 0 else PatternFill("solid", fgColor=_CINZA_CLARO)
        valores = [
            nota.chave, nota.num_doc, nota.serie, nota.dt_emissao,
            nota.nome_emit, nota.cnpj_emit,
            nota.nome_dest, nota.cnpj_dest,
            nota.vl_total, nota.situacao,
        ]
        for c, v in enumerate(valores, 1):
            cell = ws.cell(r, c, v)
            cell.fill = fill

    _ajustar_colunas(ws)


def exportar_xlsx(
    caminho: str,
    empresa: EmpresaSped,
    resultado: ResultadoConciliacao,
):
    wb = openpyxl.Workbook()
    _aba_resumo(wb, empresa, resultado)
    _aba_entradas(wb, resultado)
    _aba_saidas(wb, resultado)
    wb.save(caminho)
